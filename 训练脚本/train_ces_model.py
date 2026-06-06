# -*- coding: utf-8 -*-
"""训练本地 CES 多标签文本分类模型。"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

import joblib
import numpy as np
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import SGDClassifier
from sklearn.metrics import precision_recall_fscore_support
from sklearn.model_selection import train_test_split
from sklearn.multiclass import OneVsRestClassifier
from sklearn.pipeline import Pipeline


RANDOM_SEED = 42
ROOT = Path(__file__).resolve().parent
SOURCE_DIR = ROOT / "training_sources" / "unsplit"
COMMENT_FILE = SOURCE_DIR / "评论汇总.xlsx"
DETAIL_FILE = SOURCE_DIR / "分类明细.xlsx"
INDICATOR_FILE = SOURCE_DIR / "指标说明.xlsx"
SOURCE_STATS_FILE = SOURCE_DIR / "来源统计.xlsx"
SUMMARY_FILE = SOURCE_DIR / "统计汇总.xlsx"
UNMATCHED_FILE = SOURCE_DIR / "未命中评论.xlsx"

MODELS_DIR = ROOT / "models"
MODEL_FILE = MODELS_DIR / "ces_text_classifier.joblib"
SCHEMA_FILE = MODELS_DIR / "ces_label_schema.json"
REPORT_FILE = MODELS_DIR / "ces_model_report.md"
VALIDATION_CSV = MODELS_DIR / "ces_model_label_metrics.csv"

TRAINING_JSON = ROOT / "ces_training_data.json"
TAXONOMY_PY = ROOT / "ces_taxonomy.py"
HTML_FILES = [ROOT / "ces-local.html", ROOT / "ces-standalone.html"]
SW_FILE = ROOT / "sw.js"

REQUIRED_COLUMNS = {
    "comments": [
        "评论ID",
        "来源文件",
        "来源平台",
        "源工作表",
        "源行号",
        "景点名",
        "评分",
        "评论时间",
        "原评论全文",
        "命中指标数",
        "命中指标",
        "命中子项",
        "命中内容",
        "命中分句摘要",
    ],
    "details": [
        "评论ID",
        "来源文件",
        "来源平台",
        "源工作表",
        "源行号",
        "景点名",
        "评分",
        "评论时间",
        "原评论全文",
        "分句序号",
        "评论分句",
        "指标编号",
        "指标名称",
        "命中子项",
        "命中内容",
        "命中词数",
    ],
    "indicators": ["指标编号", "指标名称", "指标说明", "子项", "关键词"],
}


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_sheet(path: Path, required_columns: list[str] | None = None) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = [cell_text(value) for value in next(rows)]
    if required_columns:
        missing = [name for name in required_columns if name not in header]
        if missing:
            raise ValueError(f"{path.name} 缺少列：{missing}")
    records: list[dict[str, str]] = []
    for raw_row in rows:
        record = {
            header[index]: cell_text(raw_row[index]) if index < len(raw_row) else ""
            for index in range(len(header))
        }
        if any(record.values()):
            records.append(record)
    return records


def split_terms(value: str) -> list[str]:
    terms: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[、,，;；\r\n]+", value or ""):
        term = part.strip()
        if term and term not in seen:
            seen.add(term)
            terms.append(term)
    return terms


def split_semicolon(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[；;\r\n]+", value or "") if part.strip()]


def parse_numbered_label(value: str) -> tuple[str, str]:
    match = re.match(r"^(\d+)\.(.+)$", value.strip())
    if not match:
        raise ValueError(f"标签格式错误：{value}")
    return str(int(match.group(1))), match.group(2).strip()


def sub_key(code: str, sub_index: int) -> str:
    return f"sub_c{int(code):02d}_s{sub_index:02d}"


def main_key(code: str) -> str:
    return f"main_c{int(code):02d}"


def build_taxonomy(indicator_rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    grouped: OrderedDict[str, dict[str, Any]] = OrderedDict()
    main_labels: list[dict[str, Any]] = []
    sub_labels: list[dict[str, Any]] = []
    for row in indicator_rows:
        code = str(int(row["指标编号"]))
        category = row["指标名称"]
        subcategory = row["子项"]
        keywords = split_terms(row["关键词"])
        if not category or not subcategory or not keywords:
            raise ValueError(f"指标说明存在空类别、空子项或空关键词：{row}")
        item = grouped.setdefault(
            code,
            {
                "code": code,
                "category": category,
                "description": row["指标说明"],
                "subcategories": [],
            },
        )
        if item["category"] != category:
            raise ValueError(f"指标编号 {code} 对应多个名称：{item['category']} / {category}")
        key = sub_key(code, len(item["subcategories"]) + 1)
        item["subcategories"].append({"key": key, "name": subcategory, "keywords": keywords})
    taxonomy = list(grouped.values())
    taxonomy.sort(key=lambda item: int(item["code"]))
    if {item["code"] for item in taxonomy} != {str(index) for index in range(1, 13)}:
        raise ValueError("一级类别不是 1-12")
    if sum(len(item["subcategories"]) for item in taxonomy) != 51:
        raise ValueError("二级子类不是 51 个")
    for item in taxonomy:
        main_labels.append(
            {
                "id": main_key(item["code"]),
                "level": "main",
                "code": item["code"],
                "name": item["category"],
            }
        )
        for sub in item["subcategories"]:
            sub_labels.append(
                {
                    "id": sub["key"],
                    "level": "sub",
                    "code": item["code"],
                    "category": item["category"],
                    "name": sub["name"],
                }
            )
    return taxonomy, main_labels, sub_labels


def labels_from_comment(row: dict[str, str], sub_name_to_key: dict[tuple[str, str], str]) -> tuple[list[str], list[str]]:
    main_ids: list[str] = []
    sub_ids: list[str] = []
    for raw_label in split_semicolon(row["命中指标"]):
        code, _name = parse_numbered_label(raw_label)
        main_ids.append(main_key(code))
    for raw_label in split_semicolon(row["命中子项"]):
        code, name = parse_numbered_label(raw_label)
        key = sub_name_to_key.get((code, name))
        if not key:
            raise ValueError(f"评论 {row['评论ID']} 存在未定义二级子类：{raw_label}")
        sub_ids.append(key)
    return sorted(set(main_ids)), sorted(set(sub_ids))


def build_samples(
    comment_rows: list[dict[str, str]],
    sub_name_to_key: dict[tuple[str, str], str],
) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    for row in comment_rows:
        main_ids, sub_ids = labels_from_comment(row, sub_name_to_key)
        samples.append(
            {
                "comment_id": int(row["评论ID"]),
                "source_file": row["来源文件"],
                "source_platform": row["来源平台"],
                "source_sheet": row["源工作表"],
                "source_row": int(row["源行号"]) if row["源行号"].isdigit() else row["源行号"],
                "place": row["景点名"],
                "rating": row["评分"],
                "comment_time": row["评论时间"],
                "comment": row["原评论全文"],
                "main_label_ids": main_ids,
                "sub_label_ids": sub_ids,
                "matched_main": split_semicolon(row["命中指标"]),
                "matched_subcategories": split_semicolon(row["命中子项"]),
                "matched_terms": split_terms(row["命中内容"]),
                "matched_sentence_summary": split_semicolon(row["命中分句摘要"]),
            }
        )
    return samples


def build_details(detail_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    for row in detail_rows:
        details.append(
            {
                "comment_id": int(row["评论ID"]),
                "source_file": row["来源文件"],
                "source_platform": row["来源平台"],
                "source_sheet": row["源工作表"],
                "source_row": int(row["源行号"]) if row["源行号"].isdigit() else row["源行号"],
                "place": row["景点名"],
                "rating": row["评分"],
                "comment_time": row["评论时间"],
                "comment": row["原评论全文"],
                "sentence_index": int(row["分句序号"]) if row["分句序号"].isdigit() else row["分句序号"],
                "sentence": row["评论分句"],
                "code": str(int(row["指标编号"])),
                "category": row["指标名称"],
                "subcategory": row["命中子项"],
                "matched_terms": split_terms(row["命中内容"]),
                "matched_term_count": int(row["命中词数"]) if row["命中词数"].isdigit() else 0,
            }
        )
    return details


def build_label_matrix(samples: list[dict[str, Any]], labels: list[dict[str, Any]]) -> np.ndarray:
    index = {label["id"]: idx for idx, label in enumerate(labels)}
    y = np.zeros((len(samples), len(labels)), dtype=np.int8)
    for row_index, sample in enumerate(samples):
        for label_id in sample["main_label_ids"] + sample["sub_label_ids"]:
            y[row_index, index[label_id]] = 1
    return y


def validate_label_coverage(y: np.ndarray, labels: list[dict[str, Any]], name: str) -> None:
    counts = y.sum(axis=0)
    missing = [labels[index]["id"] for index, count in enumerate(counts) if int(count) == 0]
    if missing:
        raise ValueError(f"{name} 存在无样本标签：{missing}")


def choose_thresholds(y_true: np.ndarray, probabilities: np.ndarray) -> list[float]:
    thresholds: list[float] = []
    for label_index in range(y_true.shape[1]):
        best_threshold = 0.5
        best_f1 = -1.0
        candidates = np.unique(probabilities[:, label_index])
        candidates = np.unique(np.concatenate([candidates, np.array([0.5])]))
        for threshold in candidates:
            pred = probabilities[:, label_index] >= threshold
            _p, _r, f1, _s = precision_recall_fscore_support(
                y_true[:, label_index],
                pred,
                average="binary",
                zero_division=0,
            )
            if f1 > best_f1 or (f1 == best_f1 and abs(float(threshold) - 0.5) < abs(best_threshold - 0.5)):
                best_f1 = float(f1)
                best_threshold = float(threshold)
        thresholds.append(best_threshold)
    return thresholds


def metric_block(y_true: np.ndarray, y_pred: np.ndarray) -> dict[str, float]:
    micro = precision_recall_fscore_support(y_true, y_pred, average="micro", zero_division=0)
    macro = precision_recall_fscore_support(y_true, y_pred, average="macro", zero_division=0)
    return {
        "micro_precision": float(micro[0]),
        "micro_recall": float(micro[1]),
        "micro_f1": float(micro[2]),
        "macro_precision": float(macro[0]),
        "macro_recall": float(macro[1]),
        "macro_f1": float(macro[2]),
    }


def per_label_metrics(
    labels: list[dict[str, Any]],
    y_true: np.ndarray,
    y_pred: np.ndarray,
    thresholds: list[float],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    precision, recall, f1, support = precision_recall_fscore_support(
        y_true,
        y_pred,
        average=None,
        zero_division=0,
    )
    for index, label in enumerate(labels):
        rows.append(
            {
                **label,
                "support": int(support[index]),
                "threshold": float(thresholds[index]),
                "precision": float(precision[index]),
                "recall": float(recall[index]),
                "f1": float(f1[index]),
            }
        )
    return rows


def keyword_map(taxonomy: list[dict[str, Any]]) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    for item in taxonomy:
        seen: set[str] = set()
        terms: list[str] = []
        for sub in item["subcategories"]:
            for keyword in sub["keywords"]:
                if keyword not in seen:
                    seen.add(keyword)
                    terms.append(keyword)
            result[sub["key"]] = sub["keywords"]
        result[main_key(item["code"])] = terms
    return result


def client_ces(taxonomy: list[dict[str, Any]]) -> list[dict[str, Any]]:
    keywords = keyword_map(taxonomy)
    return [
        {"name": item["category"], "code": item["code"], "keywords": keywords[main_key(item["code"])]}
        for item in taxonomy
    ]


def build_stats(
    taxonomy: list[dict[str, Any]],
    samples: list[dict[str, Any]],
    details: list[dict[str, Any]],
    label_metrics: dict[str, Any],
) -> dict[str, Any]:
    main_counter = Counter(label_id for sample in samples for label_id in sample["main_label_ids"])
    sub_counter = Counter(label_id for sample in samples for label_id in sample["sub_label_ids"])
    detail_counter = Counter(detail["code"] for detail in details)
    return {
        "dataset_id": "unsplit_20260523",
        "dataset_name": "未拆分 CES 指标评论分句分类结果",
        "category_count": len(taxonomy),
        "subcategory_count": sum(len(item["subcategories"]) for item in taxonomy),
        "keyword_count": sum(len(sub["keywords"]) for item in taxonomy for sub in item["subcategories"]),
        "sample_count": len(samples),
        "sentence_hit_count": len(details),
        "model": label_metrics,
        "counts_by_category": [
            {
                "code": item["code"],
                "category": item["category"],
                "sample_count": int(main_counter[main_key(item["code"])]),
                "sentence_hit_count": int(detail_counter[item["code"]]),
            }
            for item in taxonomy
        ],
        "counts_by_subcategory": [
            {
                "id": sub["key"],
                "code": item["code"],
                "category": item["category"],
                "subcategory": sub["name"],
                "sample_count": int(sub_counter[sub["key"]]),
            }
            for item in taxonomy
            for sub in item["subcategories"]
        ],
    }


def write_training_json(
    taxonomy: list[dict[str, Any]],
    samples: list[dict[str, Any]],
    details: list[dict[str, Any]],
    stats: dict[str, Any],
) -> None:
    unmatched_rows = read_sheet(UNMATCHED_FILE) if UNMATCHED_FILE.exists() else []
    source_stats = read_sheet(SOURCE_STATS_FILE) if SOURCE_STATS_FILE.exists() else []
    summary_rows = read_sheet(SUMMARY_FILE) if SUMMARY_FILE.exists() else []
    payload = {
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "dataset_id": "unsplit_20260523",
            "source_files": [
                str(COMMENT_FILE),
                str(DETAIL_FILE),
                str(INDICATOR_FILE),
                str(SOURCE_STATS_FILE),
                str(SUMMARY_FILE),
                str(UNMATCHED_FILE),
            ],
            "matching_policy": "exact substring match for explanations only; ML model is classification source",
        },
        "taxonomy": taxonomy,
        "samples": samples,
        "sentence_hits": details,
        "unmatched_comments": unmatched_rows,
        "source_stats": source_stats,
        "summary_rows": summary_rows,
        "stats": stats,
    }
    TRAINING_JSON.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def write_taxonomy_py(taxonomy: list[dict[str, Any]], stats: dict[str, Any]) -> None:
    content = (
        "# -*- coding: utf-8 -*-\n"
        '"""CES 12 类未拆分训练集分类库，由 train_ces_model.py 从 Excel 生成。"""\n\n'
        "CES_TAXONOMY = "
        + json.dumps(taxonomy, ensure_ascii=False, indent=2)
        + "\n\n"
        "CES_TRAINING_STATS = "
        + json.dumps(stats, ensure_ascii=False, indent=2)
        + "\n"
    )
    TAXONOMY_PY.write_text(content, encoding="utf-8")


def update_html(taxonomy: list[dict[str, Any]], stats: dict[str, Any]) -> None:
    replacement = (
        "var FULL_CES_TAXONOMY = "
        + json.dumps(taxonomy, ensure_ascii=False, indent=2)
        + ";\n\n"
        "var CES_TRAINING_STATS = "
        + json.dumps(stats, ensure_ascii=False, indent=2)
        + ";\n\n"
        "var CLIENT_CES = "
        + json.dumps(client_ces(taxonomy), ensure_ascii=False, indent=2)
        + ";\n\n"
    )
    for html_file in HTML_FILES:
        text = html_file.read_text(encoding="utf-8")
        start = text.index("var FULL_CES_TAXONOMY")
        end = text.index("var API_BASE")
        text = text[:start] + replacement + text[end:]
        html_file.write_text(text, encoding="utf-8")


def update_service_worker() -> None:
    text = SW_FILE.read_text(encoding="utf-8")
    text = re.sub(
        r'const CACHE_NAME = "ces-sentiment-[^"]+";',
        'const CACHE_NAME = "ces-sentiment-v8-ml-unsplit";',
        text,
        count=1,
    )
    SW_FILE.write_text(text, encoding="utf-8")


def write_schema(
    taxonomy: list[dict[str, Any]],
    labels: list[dict[str, Any]],
    main_labels: list[dict[str, Any]],
    sub_labels: list[dict[str, Any]],
    thresholds: list[float],
    metrics: dict[str, Any],
) -> dict[str, Any]:
    schema = {
        "model_id": "ces_text_classifier_unsplit_20260523",
        "model_type": "char_tfidf_sgd_ovr_multilabel",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "random_seed": RANDOM_SEED,
        "dataset": {
            "id": "unsplit_20260523",
            "comment_file": str(COMMENT_FILE),
            "detail_file": str(DETAIL_FILE),
            "indicator_file": str(INDICATOR_FILE),
        },
        "taxonomy": taxonomy,
        "labels": labels,
        "main_labels": main_labels,
        "sub_labels": sub_labels,
        "thresholds": {label["id"]: thresholds[index] for index, label in enumerate(labels)},
        "metrics": metrics,
    }
    SCHEMA_FILE.write_text(json.dumps(schema, ensure_ascii=False, indent=2), encoding="utf-8")
    return schema


def write_report(schema: dict[str, Any], label_rows: list[dict[str, Any]], stats: dict[str, Any]) -> None:
    metrics = schema["metrics"]
    lines = [
        "# CES 本地机器学习模型报告",
        "",
        f"- 模型：{schema['model_type']}",
        f"- 数据集：{schema['dataset']['id']}",
        f"- 评论样本数：{stats['sample_count']}",
        f"- 分句命中明细数：{stats['sentence_hit_count']}",
        f"- 一级类别数：{stats['category_count']}",
        f"- 二级子类数：{stats['subcategory_count']}",
        f"- 触发词数：{stats['keyword_count']}",
        f"- 验证集样本数：{metrics['validation_size']}",
        "",
        "## 总体指标",
        "",
        "| 标签层级 | Micro F1 | Macro F1 | Precision | Recall |",
        "|---|---:|---:|---:|---:|",
    ]
    for key, title in [("main", "一级类别"), ("sub", "二级子类"), ("all", "全部标签")]:
        item = metrics[key]
        lines.append(
            f"| {title} | {item['micro_f1']:.4f} | {item['macro_f1']:.4f} | "
            f"{item['micro_precision']:.4f} | {item['micro_recall']:.4f} |"
        )
    lines.extend(
        [
            "",
            "## 标签明细",
            "",
            "| 层级 | 编码 | 名称 | 验证样本数 | 阈值 | Precision | Recall | F1 |",
            "|---|---|---|---:|---:|---:|---:|---:|",
        ]
    )
    for row in label_rows:
        name = row["name"]
        code = row["code"]
        level = "一级" if row["level"] == "main" else "二级"
        lines.append(
            f"| {level} | {code} | {name} | {row['support']} | {row['threshold']:.4f} | "
            f"{row['precision']:.4f} | {row['recall']:.4f} | {row['f1']:.4f} |"
        )
    REPORT_FILE.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_label_csv(label_rows: list[dict[str, Any]]) -> None:
    with VALIDATION_CSV.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["id", "level", "code", "category", "name", "support", "threshold", "precision", "recall", "f1"],
        )
        writer.writeheader()
        for row in label_rows:
            writer.writerow(row)


def train_model() -> None:
    MODELS_DIR.mkdir(parents=True, exist_ok=True)

    indicator_rows = read_sheet(INDICATOR_FILE, REQUIRED_COLUMNS["indicators"])
    comment_rows = read_sheet(COMMENT_FILE, REQUIRED_COLUMNS["comments"])
    detail_rows = read_sheet(DETAIL_FILE, REQUIRED_COLUMNS["details"])

    taxonomy, main_labels, sub_labels = build_taxonomy(indicator_rows)
    labels = main_labels + sub_labels
    sub_name_to_key = {
        (item["code"], sub["name"]): sub["key"]
        for item in taxonomy
        for sub in item["subcategories"]
    }
    samples = build_samples(comment_rows, sub_name_to_key)
    details = build_details(detail_rows)

    if len(samples) != 15485:
        raise ValueError(f"评论样本数不是 15485：{len(samples)}")
    if len(details) != 71575:
        raise ValueError(f"分类明细行数不是 71575：{len(details)}")
    if len(main_labels) != 12 or len(sub_labels) != 51:
        raise ValueError(f"标签数量错误：一级 {len(main_labels)}，二级 {len(sub_labels)}")

    texts = [sample["comment"] for sample in samples]
    y = build_label_matrix(samples, labels)
    validate_label_coverage(y[:, : len(main_labels)], main_labels, "一级类别")
    validate_label_coverage(y[:, len(main_labels) :], sub_labels, "二级子类")

    indices = np.arange(len(samples))
    train_idx, valid_idx = train_test_split(indices, test_size=0.25, random_state=RANDOM_SEED, shuffle=True)
    y_train = y[train_idx]
    y_valid = y[valid_idx]
    validate_label_coverage(y_train[:, : len(main_labels)], main_labels, "训练集一级类别")
    validate_label_coverage(y_train[:, len(main_labels) :], sub_labels, "训练集二级子类")
    validate_label_coverage(y_valid[:, : len(main_labels)], main_labels, "验证集一级类别")
    validate_label_coverage(y_valid[:, len(main_labels) :], sub_labels, "验证集二级子类")

    pipeline = Pipeline(
        steps=[
            (
                "tfidf",
                TfidfVectorizer(
                    analyzer="char",
                    ngram_range=(2, 5),
                    min_df=2,
                    max_features=120000,
                    sublinear_tf=True,
                ),
            ),
            (
                "clf",
                OneVsRestClassifier(
                    SGDClassifier(
                        loss="log",
                        penalty="l2",
                        alpha=1e-5,
                        max_iter=1000,
                        tol=1e-3,
                        class_weight="balanced",
                        random_state=RANDOM_SEED,
                    ),
                    n_jobs=1,
                ),
            ),
        ]
    )
    pipeline.fit([texts[index] for index in train_idx], y_train)
    probabilities = pipeline.predict_proba([texts[index] for index in valid_idx])
    thresholds = choose_thresholds(y_valid, probabilities)
    y_pred = probabilities >= np.array(thresholds)

    main_slice = slice(0, len(main_labels))
    sub_slice = slice(len(main_labels), len(labels))
    metrics = {
        "train_size": int(len(train_idx)),
        "validation_size": int(len(valid_idx)),
        "main": metric_block(y_valid[:, main_slice], y_pred[:, main_slice]),
        "sub": metric_block(y_valid[:, sub_slice], y_pred[:, sub_slice]),
        "all": metric_block(y_valid, y_pred),
    }
    label_rows = per_label_metrics(labels, y_valid, y_pred, thresholds)
    schema = write_schema(taxonomy, labels, main_labels, sub_labels, thresholds, metrics)
    stats = build_stats(taxonomy, samples, details, metrics)

    artifact = {
        "pipeline": pipeline,
        "schema": schema,
        "keyword_map": keyword_map(taxonomy),
    }
    joblib.dump(artifact, MODEL_FILE)
    # 立刻重载，避免生成了不能用的模型文件。
    reloaded = joblib.load(MODEL_FILE)
    test_prob = reloaded["pipeline"].predict_proba([texts[0]])
    if test_prob.shape != (1, len(labels)):
        raise ValueError("模型重载后输出维度错误")

    write_training_json(taxonomy, samples, details, stats)
    write_taxonomy_py(taxonomy, stats)
    update_html(taxonomy, stats)
    update_service_worker()
    write_report(schema, label_rows, stats)
    write_label_csv(label_rows)

    print(
        json.dumps(
            {
                "model_file": str(MODEL_FILE),
                "schema_file": str(SCHEMA_FILE),
                "report_file": str(REPORT_FILE),
                "main_micro_f1": metrics["main"]["micro_f1"],
                "sub_micro_f1": metrics["sub"]["micro_f1"],
                "all_micro_f1": metrics["all"]["micro_f1"],
                "sample_count": len(samples),
                "sentence_hit_count": len(details),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    train_model()
