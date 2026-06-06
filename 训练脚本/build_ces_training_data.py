# -*- coding: utf-8 -*-
"""从 Excel 训练集生成 CES 本地数据与前后端内置词典。"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE_DIR = ROOT.parents[2] / "训练集" / "GPT拆分"
CLASSIFY_FILE = SOURCE_DIR / "4评论分类结果.xlsx"
HITS_FILE = SOURCE_DIR / "5命中明细.xlsx"
DICTIONARY_FILE = SOURCE_DIR / "6分类规则词典.xlsx"

TRAINING_JSON = ROOT / "ces_training_data.json"
VALIDATION_MD = ROOT / "ces_training_validation.md"
MISMATCH_CSV = ROOT / "ces_training_mismatches.csv"
TAXONOMY_PY = ROOT / "ces_taxonomy.py"
HTML_FILES = [ROOT / "ces-local.html", ROOT / "ces-standalone.html"]
SW_FILE = ROOT / "sw.js"


REQUIRED_COLUMNS = {
    "classify": [
        "序号",
        "来源文件",
        "来源工作表",
        "原行号",
        "平台",
        "景点/对象",
        "用户",
        "评分",
        "评论时间",
        "评论内容",
        "命中指标数",
        "命中指标编码",
        "命中指标名称",
        "命中二级类别",
        "触发词",
        "命中原文片段",
        "是否多标签",
        "备注",
    ],
    "hits": [
        "序号",
        "来源文件",
        "来源工作表",
        "原行号",
        "平台",
        "景点/对象",
        "用户",
        "评分",
        "评论时间",
        "评论预览",
        "指标编码",
        "指标名称",
        "命中二级类别",
        "触发词",
        "命中原文片段",
    ],
    "dictionary": ["指标编码", "指标名称", "二级类别", "触发词"],
}


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_sheet(path: Path, required: list[str]) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header = [cell_text(value) for value in next(rows)]
    missing = [name for name in required if name not in header]
    if missing:
        raise ValueError(f"{path.name} 缺少列：{missing}")
    records: list[dict[str, str]] = []
    for raw_row in rows:
        record = {
            header[index]: cell_text(raw_row[index]) if index < len(raw_row) else ""
            for index in range(len(header))
        }
        if any(record.values()):
            records.append(record)
    return records


def parse_codes(value: str) -> list[str]:
    codes: list[str] = []
    for part in re.split(r"[；;、,，\s]+", value):
        part = part.strip()
        if not part:
            continue
        match = re.match(r"(\d+)", part)
        if match:
            codes.append(str(int(match.group(1))))
    return sorted(set(codes), key=lambda item: int(item))


def split_terms(value: str) -> list[str]:
    seen: set[str] = set()
    terms: list[str] = []
    for part in re.split(r"[、,，;；\r\n]+", value):
        term = part.strip()
        if term and term not in seen:
            seen.add(term)
            terms.append(term)
    return terms


def split_semicolon(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[；;\r\n]+", value) if part.strip()]


def stable_subcategory_key(code: str, sub_index: int) -> str:
    return f"c{int(code):02d}_s{sub_index:02d}"


def build_taxonomy(dictionary_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    grouped: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in dictionary_rows:
        code = str(int(row["指标编码"]))
        category = row["指标名称"]
        subcategory = row["二级类别"]
        keywords = split_terms(row["触发词"])
        if not category or not subcategory or not keywords:
            raise ValueError(f"分类规则词典存在空类别、空子类或空触发词：{row}")
        item = grouped.setdefault(
            code,
            {
                "code": code,
                "category": category,
                "description": "",
                "subcategories": [],
            },
        )
        if item["category"] != category:
            raise ValueError(f"指标编码 {code} 对应多个名称：{item['category']} / {category}")
        item["subcategories"].append(
            {
                "key": stable_subcategory_key(code, len(item["subcategories"]) + 1),
                "name": subcategory,
                "keywords": keywords,
            }
        )
    taxonomy = list(grouped.values())
    expected_codes = {str(index) for index in range(1, 13)}
    actual_codes = {item["code"] for item in taxonomy}
    if actual_codes != expected_codes:
        raise ValueError(f"分类编码不是 1-12：{sorted(actual_codes, key=int)}")
    taxonomy.sort(key=lambda item: int(item["code"]))
    subcategory_count = sum(len(item["subcategories"]) for item in taxonomy)
    if subcategory_count != 51:
        raise ValueError(f"二级类别数量不是 51：{subcategory_count}")
    return taxonomy


def category_name_map(taxonomy: list[dict[str, Any]]) -> dict[str, str]:
    return {item["code"]: item["category"] for item in taxonomy}


def validate_names(
    classify_rows: list[dict[str, str]],
    hit_rows: list[dict[str, str]],
    names: dict[str, str],
) -> None:
    expected_codes = set(names)
    for row in hit_rows:
        code = str(int(row["指标编码"]))
        if code not in expected_codes:
            raise ValueError(f"命中明细存在 1-12 之外编码：{code}")
        if row["指标名称"] != names[code]:
            raise ValueError(f"命中明细类别名不一致：编码 {code}，{row['指标名称']} != {names[code]}")
    for row in classify_rows:
        codes = set(parse_codes(row["命中指标编码"]))
        extra_codes = codes - expected_codes
        if extra_codes:
            raise ValueError(f"评论分类结果存在 1-12 之外编码：{sorted(extra_codes, key=int)}")


def keywords_by_code(taxonomy: list[dict[str, Any]]) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    for category in taxonomy:
        seen: set[str] = set()
        keywords: list[str] = []
        for subcategory in category["subcategories"]:
            for keyword in subcategory["keywords"]:
                if keyword not in seen:
                    seen.add(keyword)
                    keywords.append(keyword)
        result[category["code"]] = keywords
    return result


def exact_match_codes(text: str, keyword_map: dict[str, list[str]]) -> list[str]:
    matched: list[str] = []
    for code, keywords in keyword_map.items():
        if any(keyword in text for keyword in keywords):
            matched.append(code)
    return sorted(matched, key=int)


def exact_match_keywords(text: str, keyword_map: dict[str, list[str]]) -> dict[str, list[str]]:
    matched: dict[str, list[str]] = {}
    for code, keywords in keyword_map.items():
        hits = [keyword for keyword in keywords if keyword in text]
        if hits:
            matched[code] = hits
    return matched


def build_samples(classify_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    for row in classify_rows:
        codes = parse_codes(row["命中指标编码"])
        samples.append(
            {
                "sample_id": int(row["序号"]),
                "source_file": row["来源文件"],
                "source_sheet": row["来源工作表"],
                "source_row": int(row["原行号"]) if row["原行号"].isdigit() else row["原行号"],
                "platform": row["平台"],
                "place": row["景点/对象"],
                "user": row["用户"],
                "rating": row["评分"],
                "comment_time": row["评论时间"],
                "comment": row["评论内容"],
                "matched_code_count": int(row["命中指标数"]) if row["命中指标数"].isdigit() else len(codes),
                "matched_codes": codes,
                "matched_categories": split_semicolon(row["命中指标名称"]),
                "matched_subcategories": split_semicolon(row["命中二级类别"]),
                "trigger_terms": split_semicolon(row["触发词"]),
                "matched_fragments": split_semicolon(row["命中原文片段"]),
                "multi_label": row["是否多标签"] == "是",
                "note": row["备注"],
            }
        )
    return samples


def build_hits(hit_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    hits: list[dict[str, Any]] = []
    for row in hit_rows:
        hits.append(
            {
                "sample_id": int(row["序号"]),
                "source_file": row["来源文件"],
                "source_sheet": row["来源工作表"],
                "source_row": int(row["原行号"]) if row["原行号"].isdigit() else row["原行号"],
                "platform": row["平台"],
                "place": row["景点/对象"],
                "user": row["用户"],
                "rating": row["评分"],
                "comment_time": row["评论时间"],
                "comment_preview": row["评论预览"],
                "code": str(int(row["指标编码"])),
                "category": row["指标名称"],
                "subcategory": row["命中二级类别"],
                "trigger_term": row["触发词"],
                "matched_fragment": row["命中原文片段"],
            }
        )
    return hits


def validate_exact_matches(
    samples: list[dict[str, Any]],
    keyword_map: dict[str, list[str]],
) -> tuple[dict[str, int], list[dict[str, Any]]]:
    mismatches: list[dict[str, Any]] = []
    exact_count = 0
    for sample in samples:
        expected = sample["matched_codes"]
        actual = exact_match_codes(sample["comment"], keyword_map)
        if actual == expected:
            exact_count += 1
            continue
        expected_set = set(expected)
        actual_set = set(actual)
        mismatches.append(
            {
                "sample_id": sample["sample_id"],
                "expected_codes": "；".join(expected),
                "actual_codes": "；".join(actual),
                "missing_codes": "；".join(sorted(expected_set - actual_set, key=int)),
                "extra_codes": "；".join(sorted(actual_set - expected_set, key=int)),
                "comment_preview": sample["comment"][:160],
            }
        )
    stats = {
        "total_samples": len(samples),
        "exact_match_count": exact_count,
        "mismatch_count": len(mismatches),
    }
    return stats, mismatches


def build_stats(
    taxonomy: list[dict[str, Any]],
    samples: list[dict[str, Any]],
    hits: list[dict[str, Any]],
    validation: dict[str, int],
) -> dict[str, Any]:
    hit_counter = Counter(hit["code"] for hit in hits)
    sample_counter = Counter(code for sample in samples for code in sample["matched_codes"])
    return {
        "category_count": len(taxonomy),
        "subcategory_count": sum(len(item["subcategories"]) for item in taxonomy),
        "keyword_count": sum(
            len(subcategory["keywords"])
            for item in taxonomy
            for subcategory in item["subcategories"]
        ),
        "sample_count": len(samples),
        "hit_count": len(hits),
        "hit_counts_by_category": [
            {
                "code": item["code"],
                "category": item["category"],
                "hit_count": hit_counter[item["code"]],
                "sample_count": sample_counter[item["code"]],
            }
            for item in taxonomy
        ],
        "validation": validation,
    }


def client_ces(taxonomy: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "name": item["category"],
            "code": item["code"],
            "keywords": keywords_by_code([item])[item["code"]],
        }
        for item in taxonomy
    ]


def write_training_json(
    taxonomy: list[dict[str, Any]],
    samples: list[dict[str, Any]],
    hits: list[dict[str, Any]],
    stats: dict[str, Any],
) -> None:
    payload = {
        "metadata": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "source_files": [
                str(CLASSIFY_FILE),
                str(HITS_FILE),
                str(DICTIONARY_FILE),
            ],
            "matching_policy": "exact substring match; no fuzzy matching; no generated keywords",
        },
        "taxonomy": taxonomy,
        "samples": samples,
        "hits": hits,
        "stats": stats,
    }
    TRAINING_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


def write_taxonomy_py(taxonomy: list[dict[str, Any]], stats: dict[str, Any]) -> None:
    content = (
        "# -*- coding: utf-8 -*-\n"
        '"""CES 12 类训练集分类库，由 build_ces_training_data.py 从 Excel 生成。"""\n\n'
        "CES_TAXONOMY = "
        + json.dumps(taxonomy, ensure_ascii=False, indent=2)
        + "\n\n"
        "CES_TRAINING_STATS = "
        + json.dumps(stats, ensure_ascii=False, indent=2)
        + "\n"
    )
    TAXONOMY_PY.write_text(content, encoding="utf-8")


def update_html(taxonomy: list[dict[str, Any]], stats: dict[str, Any]) -> None:
    replacement = (
        "var FULL_CES_TAXONOMY = "
        + json.dumps(taxonomy, ensure_ascii=False, indent=2)
        + ";\n\n"
        "var CES_TRAINING_STATS = "
        + json.dumps(stats, ensure_ascii=False, indent=2)
        + ";\n\n"
        "var CLIENT_CES = "
        + json.dumps(client_ces(taxonomy), ensure_ascii=False, indent=2)
        + ";\n\n"
    )
    for html_file in HTML_FILES:
        text = html_file.read_text(encoding="utf-8")
        start = text.index("var FULL_CES_TAXONOMY")
        end = text.index("var API_BASE")
        text = text[:start] + replacement + text[end:]
        text = re.sub(
            r'(<textarea id="customDictionaryInput" rows="6">)[\s\S]*?(</textarea>)',
            r"\1\2",
            text,
            count=1,
        )
        html_file.write_text(text, encoding="utf-8")


def update_service_worker() -> None:
    text = SW_FILE.read_text(encoding="utf-8")
    text = re.sub(
        r'const CACHE_NAME = "ces-sentiment-[^"]+";',
        'const CACHE_NAME = "ces-sentiment-v7-excel12";',
        text,
        count=1,
    )
    assets = (
        'const ASSETS = [\n'
        '  "./",\n'
        '  "./ces-standalone.html",\n'
        '  "./ces-local.html",\n'
        '  "./manifest.webmanifest",\n'
        '  "./icon.svg",\n'
        "];"
    )
    text = re.sub(r"const ASSETS = \[[\s\S]*?\];", assets, text, count=1)
    SW_FILE.write_text(text, encoding="utf-8")


def write_validation_report(
    taxonomy: list[dict[str, Any]],
    stats: dict[str, Any],
    mismatches: list[dict[str, Any]],
) -> None:
    lines = [
        "# CES 训练集接入校验报告",
        "",
        f"- 分类数量：{stats['category_count']}",
        f"- 二级类别数量：{stats['subcategory_count']}",
        f"- 触发词数量：{stats['keyword_count']}",
        f"- 评论样本数：{stats['sample_count']}",
        f"- 命中明细数：{stats['hit_count']}",
        f"- 完全一致样本数：{stats['validation']['exact_match_count']}",
        f"- 不一致样本数：{stats['validation']['mismatch_count']}",
        "",
        "## 12 类命中统计",
        "",
        "| 编码 | 类别 | 样本数 | 命中明细数 |",
        "|---|---|---:|---:|",
    ]
    for item in stats["hit_counts_by_category"]:
        lines.append(
            f"| {item['code']} | {item['category']} | {item['sample_count']} | {item['hit_count']} |"
        )
    lines.extend(["", "## 二级类别", ""])
    for item in taxonomy:
        sub_names = "、".join(subcategory["name"] for subcategory in item["subcategories"])
        lines.append(f"- {item['code']} {item['category']}：{sub_names}")
    lines.extend(["", "## 精确命中不一致样本", ""])
    if not mismatches:
        lines.append("- 无。")
    else:
        lines.append(f"- 已保存完整明细：`{MISMATCH_CSV.name}`。")
        lines.append("- 以下只列前 20 条：")
        lines.append("")
        lines.append("| 样本 | Excel编码 | 新词典编码 | 缺少 | 多出 | 评论预览 |")
        lines.append("|---:|---|---|---|---|---|")
        for row in mismatches[:20]:
            preview = row["comment_preview"].replace("|", "｜").replace("\n", " ")
            lines.append(
                f"| {row['sample_id']} | {row['expected_codes']} | {row['actual_codes']} | "
                f"{row['missing_codes']} | {row['extra_codes']} | {preview} |"
            )
    VALIDATION_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_mismatch_csv(mismatches: list[dict[str, Any]]) -> None:
    with MISMATCH_CSV.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "sample_id",
                "expected_codes",
                "actual_codes",
                "missing_codes",
                "extra_codes",
                "comment_preview",
            ],
        )
        writer.writeheader()
        writer.writerows(mismatches)


def main() -> None:
    classify_rows = read_sheet(CLASSIFY_FILE, REQUIRED_COLUMNS["classify"])
    hit_rows = read_sheet(HITS_FILE, REQUIRED_COLUMNS["hits"])
    dictionary_rows = read_sheet(DICTIONARY_FILE, REQUIRED_COLUMNS["dictionary"])

    taxonomy = build_taxonomy(dictionary_rows)
    names = category_name_map(taxonomy)
    validate_names(classify_rows, hit_rows, names)

    samples = build_samples(classify_rows)
    hits = build_hits(hit_rows)
    keyword_map = keywords_by_code(taxonomy)
    validation, mismatches = validate_exact_matches(samples, keyword_map)
    stats = build_stats(taxonomy, samples, hits, validation)

    write_training_json(taxonomy, samples, hits, stats)
    write_taxonomy_py(taxonomy, stats)
    update_html(taxonomy, stats)
    update_service_worker()
    write_mismatch_csv(mismatches)
    write_validation_report(taxonomy, stats, mismatches)

    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
